import streamlit as st
import pandas as pd
import numpy as np
from workalendar.america import Brazil
import openpyxl
import os
from utils.functions.date_functions import *
from utils.user import *
from utils.queries import GET_PERMISSIONS, GET_USERNAME, get_casas_validas



def config_permissoes_user():
    email = st.session_state.get("user_email", "Usuário desconhecido")
    dfpermissao = GET_PERMISSIONS(email)
    if dfpermissao.empty: # Não está no EPM
        permissao = ["Gazit"]
        nomeUser = ""
    else: # Está no EPM
        permissao = dfpermissao["Permissao"].tolist()
        nomeUser = GET_USERNAME(email)
        nomeUser = " ".join(nomeUser["Nome"].tolist())
    return permissao, nomeUser, email


def mostrar_menu_permissoes_eventos(permissoes):
    if "Dev Dash Eventos" in permissoes:
        st.sidebar.markdown("## Eventos")
        st.sidebar.page_link("pages/1_Calendário_de_Eventos.py", label=":calendar: Calendário de Eventos")
        st.sidebar.page_link("pages/2_KPIs_Conversao_Eventos_Priceless.py", label="📈 KPI's de Vendas - Conversão de Eventos")
        st.sidebar.page_link("pages/3_Acompanhamento_de_Comissão.py", label="📊 KPI's de Vendas - Cálculo da Comissão de Eventos")
        st.sidebar.page_link("pages/9_KPIs_Historico_Clientes_Eventos.py", label=":busts_in_silhouette: KPI's de Vendas - Histórico e Recorrência de Clientes")
        st.sidebar.page_link("pages/4_Faturamento_Bruto_de_Eventos.py", label=":moneybag: Faturamento Bruto de Eventos")
        st.sidebar.page_link("pages/5_Conciliação_de_Parcelas_Eventos.py", label=":left_right_arrow: Conciliação de Parcelas de Eventos")
        st.sidebar.page_link("pages/8_Auditoria_de_Eventos_Preenchimento_Lancamentos.py", label=":receipt: Auditoria de Eventos - Preenchimento dos Lançamentos")
        st.sidebar.page_link("pages/6_Informações_de_Eventos.py", label="🔎 Informações de Eventos")
        # st.sidebar.page_link("pages/10_Regras_de_Eventos.py", label=":clipboard: Regras de Eventos")
        st.sidebar.page_link("pages/7_Gazit.py", label=":shopping_bags: Gazit")
        st.sidebar.page_link("pages/10_Calendário_Gazit.py", label=":calendar: Calendário de Eventos - Gazit")
        st.sidebar.page_link("pages/11_Calendário_de_Eventos_Confirmados.py", label=":calendar: Calendário de Eventos Confirmados")
    elif "Liderança Comercial Dash Eventos" in permissoes:
        st.sidebar.title("## Eventos")
        st.sidebar.page_link("pages/1_Calendário_de_Eventos.py", label=":calendar: Calendário de Eventos")
        st.sidebar.page_link("pages/2_KPIs_Conversao_Eventos_Priceless.py", label="📈 KPI's de Vendas - Conversão de Eventos")
        st.sidebar.page_link("pages/3_Acompanhamento_de_Comissão.py", label="📊 KPI's de Vendas - Cálculo da Comissão de Eventos")
        st.sidebar.page_link("pages/9_KPIs_Historico_Clientes_Eventos.py", label=":busts_in_silhouette: KPI's de Vendas - Histórico e Recorrência de Clientes")
        st.sidebar.page_link("pages/4_Faturamento_Bruto_de_Eventos.py", label=":moneybag: Faturamento Bruto de Eventos")
        st.sidebar.page_link("pages/5_Conciliação_de_Parcelas_Eventos.py", label=":left_right_arrow: Conciliação de Parcelas de Eventos")
        st.sidebar.page_link("pages/8_Auditoria_de_Eventos_Preenchimento_Lancamentos.py", label=":receipt: Auditoria de Eventos - Preenchimento dos Lançamentos")
        st.sidebar.page_link("pages/6_Informações_de_Eventos.py", label="🔎 Informações de Eventos")
        # st.sidebar.page_link("pages/10_Regras_de_Eventos.py", label=":clipboard: Regras de Eventos")
        st.sidebar.page_link("pages/7_Gazit.py", label=":shopping_bags: Gazit")
        st.sidebar.page_link("pages/10_Calendário_Gazit.py", label=":calendar: Calendário de Eventos - Gazit")
        st.sidebar.page_link("pages/11_Calendário_de_Eventos_Confirmados.py", label=":calendar: Calendário de Eventos Confirmados")
    elif "Admin Dash Eventos" in permissoes:
        st.sidebar.title("## Eventos")
        st.sidebar.page_link("pages/1_Calendário_de_Eventos.py", label=":calendar: Calendário de Eventos")
        st.sidebar.page_link("pages/2_KPIs_Conversao_Eventos_Priceless.py", label="📈 KPI's de Vendas - Conversão de Eventos")
        st.sidebar.page_link("pages/3_Acompanhamento_de_Comissão.py", label="📊 KPI's de Vendas - Cálculo da Comissão de Eventos")
        st.sidebar.page_link("pages/9_KPIs_Historico_Clientes_Eventos.py", label=":busts_in_silhouette: KPI's de Vendas - Histórico e Recorrência de Clientes")
        st.sidebar.page_link("pages/4_Faturamento_Bruto_de_Eventos.py", label=":moneybag: Faturamento Bruto de Eventos")
        st.sidebar.page_link("pages/5_Conciliação_de_Parcelas_Eventos.py", label=":left_right_arrow: Conciliação de Parcelas de Eventos")
        st.sidebar.page_link("pages/8_Auditoria_de_Eventos_Preenchimento_Lancamentos.py", label=":receipt: Auditoria de Eventos - Preenchimento dos Lançamentos")
        st.sidebar.page_link("pages/6_Informações_de_Eventos.py", label="🔎 Informações de Eventos")
        # st.sidebar.page_link("pages/10_Regras_de_Eventos.py", label=":clipboard: Regras de Eventos")
        st.sidebar.page_link("pages/7_Gazit.py", label=":shopping_bags: Gazit")
        st.sidebar.page_link("pages/10_Calendário_Gazit.py", label=":calendar: Calendário de Eventos - Gazit")
    elif "Dash Eventos Acesso 1" in permissoes:
        st.sidebar.title("## Eventos")
        st.sidebar.page_link("pages/1_Calendário_de_Eventos.py", label=":calendar: Calendário de Eventos")
        st.sidebar.page_link("pages/2_KPIs_Conversao_Eventos_Priceless.py", label="📈 KPI's de Vendas - Conversão de Eventos")
        st.sidebar.page_link("pages/3_Acompanhamento_de_Comissão.py", label="📊 KPI's de Vendas - Cálculo da Comissão de Eventos")
        st.sidebar.page_link("pages/9_KPIs_Historico_Clientes_Eventos.py", label=":busts_in_silhouette: KPI's de Vendas - Histórico e Recorrência de Clientes")
        st.sidebar.page_link("pages/4_Faturamento_Bruto_de_Eventos.py", label=":moneybag: Faturamento Bruto de Eventos")
        st.sidebar.page_link("pages/5_Conciliação_de_Parcelas_Eventos.py", label=":left_right_arrow: Conciliação de Parcelas de Eventos")
        st.sidebar.page_link("pages/8_Auditoria_de_Eventos_Preenchimento_Lancamentos.py", label=":receipt: Auditoria de Eventos - Preenchimento dos Lançamentos")
        st.sidebar.page_link("pages/6_Informações_de_Eventos.py", label="🔎 Informações de Eventos")
    elif "Dash Eventos Acesso 2" in permissoes:
        st.sidebar.title("## Eventos")
        st.sidebar.page_link("pages/1_Calendário_de_Eventos.py", label=":calendar: Calendário de Eventos")
        st.sidebar.page_link("pages/2_KPIs_Conversao_Eventos_Priceless.py", label="📈 KPI's de Vendas - Conversão de Eventos")
        st.sidebar.page_link("pages/9_KPIs_Historico_Clientes_Eventos.py", label=":busts_in_silhouette: KPI's de Vendas - Histórico e Recorrência de Clientes")
        st.sidebar.page_link("pages/4_Faturamento_Bruto_de_Eventos.py", label=":moneybag: Faturamento Bruto de Eventos")
        st.sidebar.page_link("pages/5_Conciliação_de_Parcelas_Eventos.py", label=":left_right_arrow: Conciliação de Parcelas de Eventos")
        st.sidebar.page_link("pages/8_Auditoria_de_Eventos_Preenchimento_Lancamentos.py", label=":receipt: Auditoria de Eventos - Preenchimento dos Lançamentos")
        st.sidebar.page_link("pages/6_Informações_de_Eventos.py", label="🔎 Informações de Eventos")
        st.sidebar.page_link("pages/7_Gazit.py", label=":shopping_bags: Gazit")
        st.sidebar.page_link("pages/10_Calendário_Gazit.py", label=":calendar: Calendário de Eventos - Gazit")
    elif "Dash Eventos Acesso 3" in permissoes:
        st.sidebar.title("## Eventos")
        st.sidebar.page_link("pages/1_Calendário_de_Eventos.py", label=":calendar: Calendário de Eventos")
        st.sidebar.page_link("pages/2_KPIs_Conversao_Eventos_Priceless.py", label="📈 KPI's de Vendas - Conversão de Eventos")
        st.sidebar.page_link("pages/9_KPIs_Historico_Clientes_Eventos.py", label=":busts_in_silhouette: KPI's de Vendas - Histórico e Recorrência de Clientes")
        st.sidebar.page_link("pages/4_Faturamento_Bruto_de_Eventos.py", label=":moneybag: Faturamento Bruto de Eventos")
        st.sidebar.page_link("pages/5_Conciliação_de_Parcelas_Eventos.py", label=":left_right_arrow: Conciliação de Parcelas de Eventos")
        st.sidebar.page_link("pages/8_Auditoria_de_Eventos_Preenchimento_Lancamentos.py", label=":receipt: Auditoria de Eventos - Preenchimento dos Lançamentos")
        st.sidebar.page_link("pages/6_Informações_de_Eventos.py", label="🔎 Informações de Eventos")
    elif "Dash Eventos Acesso 4" in permissoes:
        st.sidebar.title("## Eventos")
        st.sidebar.page_link("pages/1_Calendário_de_Eventos.py", label=":calendar: Calendário de Eventos")
        st.sidebar.page_link("pages/2_KPIs_Conversao_Eventos_Priceless.py", label="📈 KPI's de Vendas - Conversão de Eventos")
        st.sidebar.page_link("pages/9_KPIs_Historico_Clientes_Eventos.py", label=":busts_in_silhouette: KPI's de Vendas - Histórico e Recorrência de Clientes")
        st.sidebar.page_link("pages/4_Faturamento_Bruto_de_Eventos.py", label=":moneybag: Faturamento Bruto de Eventos")
        st.sidebar.page_link("pages/6_Informações_de_Eventos.py", label="🔎 Informações de Eventos")
    elif "Dash Eventos Acesso 5" in permissoes:
        st.sidebar.title("## Eventos")
        st.sidebar.page_link("pages/1_Calendário_de_Eventos.py", label=":calendar: Calendário de Eventos")
        st.sidebar.page_link("pages/9_KPIs_Historico_Clientes_Eventos.py", label=":busts_in_silhouette: KPI's de Vendas - Histórico e Recorrência de Clientes")
        st.sidebar.page_link("pages/4_Faturamento_Bruto_de_Eventos.py", label=":moneybag: Faturamento Bruto de Eventos")
        st.sidebar.page_link("pages/6_Informações_de_Eventos.py", label="🔎 Informações de Eventos")
    elif "Gazit" in permissoes:
        st.sidebar.title("## Eventos")
        st.sidebar.page_link("pages/1_Calendário_de_Eventos.py", label=":calendar: Calendário de Eventos")
        st.sidebar.page_link("pages/2_KPIs_Conversao_Eventos_Priceless.py", label="📈 KPI's de Vendas - Conversão de Eventos")
        st.sidebar.page_link("pages/9_KPIs_Historico_Clientes_Eventos.py", label=":busts_in_silhouette: KPI's de Vendas - Histórico e Recorrência de Clientes")
        st.sidebar.page_link("pages/4_Faturamento_Bruto_de_Eventos.py", label=":moneybag: Faturamento Bruto de Eventos")
        st.sidebar.page_link("pages/5_Conciliação_de_Parcelas_Eventos.py", label=":left_right_arrow: Conciliação de Parcelas de Eventos")
        st.sidebar.page_link("pages/6_Informações_de_Eventos.py", label="🔎 Informações de Eventos")
        st.sidebar.page_link("pages/7_Gazit.py", label=":shopping_bags: Gazit")
        st.sidebar.page_link("pages/10_Calendário_Gazit.py", label=":calendar: Calendário de Eventos - Gazit")
    else:
        pass


def mostrar_menu_permissoes_cmv(permissoes):
    if "Dev Dash Eventos" in permissoes:
        st.sidebar.markdown("## CMV")
        st.sidebar.page_link("pages/CMV_Teórico_-_Fichas_Técnicas.py", label=":material/rubric: CMV - Fichas Técnicas")
        st.sidebar.page_link("pages/CMV_Teórico_-_Análises.py", label=":material/shelves: CMV - Análises")

def config_sidebar():

    permissoes, user_name, email = config_permissoes_user()
    st.sidebar.header(f"Bem-vindo(a) {user_name}!")
    if st.session_state["loggedIn"]:
        mostrar_menu_permissoes_eventos(permissoes)
        mostrar_menu_permissoes_cmv(permissoes)
    else:
        st.sidebar.write("Por favor, faça login para acessar o menu.")


def filtrar_por_classe_selecionada(dataframe, classe, valores_selecionados):
    if valores_selecionados:
        dataframe = dataframe[dataframe[classe].isin(valores_selecionados)]
    return dataframe


def export_to_excel(df, sheet_name, excel_filename):
    if os.path.exists(excel_filename):
        wb = openpyxl.load_workbook(excel_filename)
    else:
        wb = openpyxl.Workbook()

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(title=sheet_name)

    # Escrever os cabeçalhos
    for col_idx, column_title in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=column_title)

    # Escrever os dados
    for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(excel_filename)


def format_brazilian(num):
    try:
        num = float(num)
        return f"{num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return num
    

def format_brazilian_without_decimal(num):
    try:
        num = float(num)
        return f"{num:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return num


def format_columns_brazilian(df, numeric_columns):
    if df is not None and not df.empty:
        df = df.copy()
        for col in numeric_columns:
            if col in df.columns:
                df[col] = df[col].apply(format_brazilian)
        return df
    else:
        return df


def format_percentage(num):
    try:
        num = float(num)
        formatted_num = f"{num * 100:,.2f}"  # Multiplica por 100 e formata
        return f"{formatted_num.replace(',', 'X').replace('.', ',').replace('X', '.')}%"  # Formata como percentual
    except (ValueError, TypeError):
        return num  # Retorna o valor original em caso de erro


def format_columns_percentage(df, numeric_columns):
    for col in numeric_columns:
        if col in df.columns:
            df[col] = df[col].apply(format_percentage)
    return df


# Dataframe filtrado pela casa:
def df_filtrar_casa(df, id_casa):
    df_filtrado = df[df["Casa"] == id_casa]
    return df_filtrado


def df_filtrar_periodo_data(df, coluna_data, data_inicio, data_fim):

    data_inicio = pd.to_datetime(data_inicio)
    data_fim = pd.to_datetime(data_fim) + pd.DateOffset(days=1)

    df = df.copy()

    df[coluna_data] = pd.to_datetime(df[coluna_data])
    df_filtrado = df.loc[
        (df[coluna_data] >= data_inicio) & (df[coluna_data] < data_fim)
    ]

    return df_filtrado


def df_filtrar_mes(df, coluna_data, mes):

    df = df.copy()

    df[coluna_data] = pd.to_datetime(df[coluna_data])

    df_filtrado = df.loc[(df[coluna_data].dt.month == int(mes))]

    return df_filtrado


def df_filtrar_ano(df, coluna_data, ano):

    df = df.copy()

    df[coluna_data] = pd.to_datetime(df[coluna_data])

    df_filtrado = df.loc[(df[coluna_data].dt.year == int(ano))]

    return df_filtrado


def escape_dolar(texto):
    return texto.replace('$', r'\$')
